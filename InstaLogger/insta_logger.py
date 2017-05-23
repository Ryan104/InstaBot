#! /usr/bin/python3
""" insta_logger.py - Logs instagram user data to an xls spreadsheet """

import time
import datetime
import json
import requests
import openpyxl

def insta_logger(user_name):
    """ log the stuff """
    user_stats = get_user_stats(get_user_json(user_name))

    print('Followers: ' + str(user_stats['num_followers']))
    print('Following: ' + str(user_stats['num_following']))
    print('Posts: ' + str(user_stats['num_posts']))
    print(user_stats['date'])

    log_to_excel(user_stats)

def get_user_json(user_name):
    """ download and return json object with user info """
    url = 'https://www.instagram.com/%s/?__a=1' % user_name
    user_json = None
    try:
        res = requests.get(url)
        res.raise_for_status()
        user_json = json.loads(res.text)
    except:
        print("exception on get_user_json")
    return user_json


def get_user_stats(user_json):
    """ parse the json data for user stats
        return dictionary with # of Followers, # of Following, Number of posts
        TODO: likes on each post, list of follwers, list of following """
    user_stats = {
        'user_name': '',
        'date': '',
        'num_followers': -1,
        'num_following': -1,
        'num_posts': -1
    }
    if user_json:
        try:
            user_stats['user_name'] = user_json['user']['username']
            user_stats['date'] = datetime.datetime.now()
            user_stats['num_followers'] = user_json['user']['followed_by']['count']
            user_stats['num_following'] = user_json['user']['follows']['count']
            user_stats['num_posts'] = user_json['user']['media']['count']
        except:
            print('exception on get_user_stats()')

    return user_stats

def log_to_excel(user_stats):
    """ saves followers, following, and number of posts to excel file """
    try:
        xfile = openpyxl.load_workbook('insta_log.xlsx')
        log_sheet = xfile.get_sheet_by_name('Sheet1')
        row = len(log_sheet['A']) + 1
        log_sheet.cell(row=row, column=1).value = user_stats['date']
        log_sheet.cell(row=row, column=2).value = user_stats['num_followers']
        log_sheet.cell(row=row, column=3).value = user_stats['num_following']
        log_sheet.cell(row=row, column=4).value = user_stats['num_posts']

        xfile.save('insta_log.xlsx')

    except:
        print('exception on log_to_excel()')

def run_at_interval(interval):
    """ run the logger at given frequency in minutes """
    while True:
        insta_logger('peace.in.solitude')
        for i in range(interval*60):
            time.sleep(1)

run_at_interval(60)
